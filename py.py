# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Color, Fill, PatternFill, GradientFill, Border, Side, Alignment, Protection

##########################
#                        #
#       written by:      #
#   Rentix Productions   #
#                        #
##########################

file_name = raw_input('File name(date): ')
file_name = file_name + '.xlsx'

#Import XLSX file
wb = openpyxl.load_workbook(file_name)

#Define Ctovet column ->>> ctovet = ctovet column
ws = wb['Sheet1']
ctovet = ws['A1']._value
listest = 'ABCDEFGHIJKLMNOPQR'
for a in listest:
	for i in range(1, 10):
		if ws[a + str(i)]._value == 'כתובת'.decode('utf-8'):
			ctovet = ws[a+str(i)]._value
			ctovet_number = i
			ctovet_letter = a

#Blackbox
def blackbox(x):
	border = Border(left=Side(border_style='thin', color='00000000'), right=Side(border_style='thin', color='00000000'), top=Side(border_style='thin', color='00000000'), bottom=Side(border_style='thin', color='00000000'), diagonal=Side(border_style='thin', color='00000000'), diagonal_direction=0, outline=Side(border_style='thin', color='00000000'), vertical=Side(border_style='thin', color='00000000'), horizontal=Side(border_style='thin', color='00000000'))
	x.border = border

#Add new columns
listest = 'BCDEFGHIJKLMNOPQRSTUVWX'
pos = -1
for i in listest:
	if ws[i+'2']._value != ws['A1']._value:
		pos += 1

city = listest[pos+1]+'2'
street = listest[pos+2]+'2'

ws[city] = 'עיר'
blackbox(ws[city])
ws[street] = 'רחוב'
blackbox(ws[street])

ws[city].font = Font(name='Calibri', b=True, sz=14)
ws[city].fill = PatternFill("solid", fgColor="a5a5a5")
ws[street].font = Font(name='Calibri', b=True, sz=14)
ws[street].fill = PatternFill("solid", fgColor="a5a5a5")

#Read & Write each row (ctovet)
cto = ctovet_number+1
while ws[ctovet_letter+str(cto)]._value != ws['A1']._value:
	line = ws[ctovet_letter+str(cto)]._value
	city = listest[pos+1]+str(cto)
	ws[city] = line[0:line.find(',')]
	street = listest[pos+2]+str(cto)
	ws[street] = line[line.find(',')+1:len(line)]
	blackbox(ws[city])
	blackbox(ws[street])
	
	cto += 1

#Define Status column ->>> status = status column
ws = wb['Sheet1']
status = ws['A1']._value
listest = 'ABCDEFGHIJKLMNOPQR'
for a in listest:
	for i in range(1, 10):
		if ws[a + str(i)]._value == 'סטטוס'.decode('utf-8'):
			status = ws[a+str(i)]._value
			status_number = i
			status_letter = a
#READ & Write each row (status)
cto = status_number+1
while ws[status_letter+str(cto)]._value != ws['A1']._value:
	city = ws[listest[pos+2]+str(cto)]
	street = ws[listest[pos+3]+str(cto)]
	if ws[status_letter+str(cto)]._value == 'פעיל'.decode('utf-8'):
		city.font = Font(name='Calibri', b=True, sz=11)
		city.fill = PatternFill("solid", fgColor="00b0f0")
		street.font = Font(name='Calibri', b=True, sz=11)
		street.fill = PatternFill("solid", fgColor="00b0f0")
	else:
		city.font = Font(name='Calibri', b=True, sz=11)
		city.fill = PatternFill("solid", fgColor="ff0000")
		street.font = Font(name='Calibri', b=True, sz=11)
		street.fill = PatternFill("solid", fgColor="ff0000")		
	
	cto += 1

print ('Success!')
print ('Check in folder for ' + file_name)
wb.save(file_name)   # use the same name if required
